# ============================================================
# FONALİST - KANTİTATİF FON YÖNETİM MOTORU (SÜRÜM V17.7)
# Özellikler: "Tarih Kayması" (Amerikan Formatı) Hatası Çözüldü,
# TEFAS Tarihsel Excel Dosyasını Doğrudan Okuma (0. Hat) Eklendi,
# Fon Adı/Alanı Eşleştirmesi Aktif, Z-Score Anomali Matrisi.
# ============================================================

import concurrent.futures
import datetime as dt
import io
import math
import re
import time
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from enum import Enum
from typing import Optional, List, Dict, Any

import numpy as np
import openpyxl
import pandas as pd
import requests
import streamlit as st
import yfinance as yf

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

try:
    from tefas import Crawler as TefasCrawler
    HAS_TEFAS_CRAWLER = True
except ImportError:
    HAS_TEFAS_CRAWLER = False

# ============================================================
# OPENPYXL 'extLst' HATASI İÇİN ÇALIŞMA ZAMANI YAMASI
# ============================================================
original_init = PatternFill.__init__
def new_init(self, *args, **kwargs):
    if 'extLst' in kwargs: del kwargs['extLst']
    original_init(self, *args, **kwargs)
PatternFill.__init__ = new_init

# ============================================================
# KANTİTATİF MOTOR SINIFLARI (DİNAMİK EŞİK & ANOMALİ)
# ============================================================

class MarketRegime(Enum):
    BULLISH = "BOGA_MOMENTUM"
    BEARISH_VOLATILE = "AYI_SAVUNMA"
    STAGNANT = "YATAY_LİKİDİTE"

class ActionSignal(Enum):
    STRONG_BUY = "GÜÇLÜ AL"
    HOLD = "ASIL LİSTE (TUT)"
    WATCH = "DÜZELTME / İZLE"
    STOP_LOSS_30 = "ZARAR KES (%30 SAT)"
    EMERGENCY_EXIT = "ACİL TASFİYE (ÇÖKÜŞ)"

@dataclass
class FundMetrics:
    code: str
    name: str
    category: str
    current_score: float
    history_scores: List[float]
    volatility: float
    aum_change_pct: float
    investor_change_pct: float
    aum_history_pct: List[float] = field(default_factory=list)
    underlying_proxy_chg: float = 0.0

class FonalistEngine:
    def __init__(self, base_threshold: float = 20.0, z_score_threshold: float = -2.5):
        self.base_threshold = base_threshold
        self.z_score_threshold = z_score_threshold

    def detect_macro_regime(self, bist30_trend: float, us10y_yield_trend: float, brent_trend: float) -> MarketRegime:
        if us10y_yield_trend > 1.5 or brent_trend > 2.0 or bist30_trend < -1.0:
            return MarketRegime.BEARISH_VOLATILE
        elif bist30_trend > 0.5 and us10y_yield_trend <= 0:
            return MarketRegime.BULLISH
        return MarketRegime.STAGNANT

    def calculate_dynamic_threshold(self, fund: FundMetrics, regime: MarketRegime) -> float:
        threshold = self.base_threshold
        if regime == MarketRegime.BEARISH_VOLATILE:
            threshold += 3.0
        elif regime == MarketRegime.BULLISH:
            threshold -= 2.0

        if fund.volatility > 2.0:
            threshold += 2.0
        elif fund.volatility < 0.5:
            threshold -= 1.0
        return max(15.0, min(threshold, 30.0))

    def detect_outflow_anomaly(self, fund: FundMetrics) -> bool:
        valid_hist = [v for v in fund.aum_history_pct if v is not None]
        if len(valid_hist) < 5:
            return fund.aum_change_pct < -15.0 or fund.investor_change_pct < -10.0
        arr = np.array(valid_hist)
        mean = np.mean(arr)
        std = np.std(arr)
        if std == 0:
            return fund.aum_change_pct < -15.0 or fund.investor_change_pct < -10.0
        z_score = (fund.aum_change_pct - mean) / std
        return z_score <= self.z_score_threshold

    def evaluate_fund(self, fund: FundMetrics, regime: MarketRegime) -> Dict:
        dynamic_cut_off = self.calculate_dynamic_threshold(fund, regime)
        is_anomaly = self.detect_outflow_anomaly(fund)

        adjusted_score = fund.current_score
        if fund.underlying_proxy_chg < -1.5:
            adjusted_score -= 4.0
        elif fund.underlying_proxy_chg > 1.5:
            adjusted_score += 3.0

        if is_anomaly:
            return {"skor": adjusted_score, "esik": dynamic_cut_off, "sinyal": ActionSignal.EMERGENCY_EXIT, "gerekce": "Z-Score Likidite Çıkış Anomalisi."}

        two_day_confirmed_bearish = (len(fund.history_scores) >= 2 and adjusted_score < dynamic_cut_off and fund.history_scores[1] < dynamic_cut_off)

        if two_day_confirmed_bearish:
            return {"skor": adjusted_score, "esik": dynamic_cut_off, "sinyal": ActionSignal.STOP_LOSS_30, "gerekce": f"2 gün üst üste eşiğin ({dynamic_cut_off:.1f}) altında kalındı."}
        
        if adjusted_score < dynamic_cut_off:
            return {"skor": adjusted_score, "esik": dynamic_cut_off, "sinyal": ActionSignal.WATCH, "gerekce": "1. Gün eşik altı. Teyit bekleniyor."}

        if adjusted_score >= 75:
            return {"skor": adjusted_score, "esik": dynamic_cut_off, "sinyal": ActionSignal.STRONG_BUY, "gerekce": "Güçlü momentum."}

        return {"skor": adjusted_score, "esik": dynamic_cut_off, "sinyal": ActionSignal.HOLD, "gerekce": "Pozisyon korunur."}

    @staticmethod
    def weights_for_horizon(yatirim_vadesi: str):
        if yatirim_vadesi == "Kısa Vade": return {"base_threshold": 24.0, "mom": 0.55, "sec": 0.25, "sent": 0.20, "proxy_scale": 1.5}
        if yatirim_vadesi == "Uzun Vade": return {"base_threshold": 17.0, "mom": 0.40, "sec": 0.45, "sent": 0.15, "proxy_scale": 0.5}
        return {"base_threshold": 20.0, "mom": 0.50, "sec": 0.35, "sent": 0.15, "proxy_scale": 1.0}

def get_live_proxy_change(fund_code: str, investment_area: str) -> float:
    proxy_map = {"AFS": "NQ=F", "THF": "XU030.IS", "KZL": "GC=F", "US10Y": "^TNX", "BRENT": "BZ=F"}
    if fund_code in proxy_map: ticker = proxy_map[fund_code]
    elif "Yabancı Hisse" in investment_area or "Yabancı Teknoloji" in investment_area: ticker = "NQ=F"
    elif "Hisse Senedi" in investment_area: ticker = "XU100.IS"
    elif "Kıymetli Maden" in investment_area: ticker = "GC=F"
    else: return 0.0
    return _fetch_yf_intraday_change(ticker)

@st.cache_data(show_spinner=False, ttl=90)
def _fetch_yf_intraday_change(ticker: str) -> float:
    try:
        data = yf.download(ticker, period="1d", interval="1m", progress=False, auto_adjust=False)
        if data is None or data.empty: return 0.0
        if isinstance(data.columns, pd.MultiIndex):
            data = data.xs(ticker, axis=1, level=1) if ticker in data.columns.get_level_values(-1) else data.droplevel(1, axis=1)
        open_price = float(data['Open'].dropna().iloc[0])
        current_price = float(data['Close'].dropna().iloc[-1])
        if open_price == 0: return 0.0
        return ((current_price / open_price) - 1.0) * 100.0
    except Exception:
        return 0.0

# ============================================================
# HTTP OTURUMU (HATA YÖNETİMİ VE YENİDEN DENEME)
# ============================================================
def build_http_session() -> requests.Session:
    session = requests.Session()
    retry = Retry(total=2, connect=2, read=2, status=2, backoff_factor=1.5, status_forcelist=(429, 500, 502, 503, 504), respect_retry_after_header=True)
    adapter = HTTPAdapter(max_retries=retry, pool_connections=10, pool_maxsize=10)
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    session.headers.update({"User-Agent": "Fonalist/17.7", "Accept": "application/json"})
    return session

HTTP = build_http_session()

# ============================================================
# STREAMLIT YAPI VE SABİTLER
# ============================================================
st.set_page_config(page_title="FONALİST Kantitatif Analiz", page_icon="📊", layout="wide")
st.title("📊 FONALİST Kantitatif Analiz Motoru")
st.caption("Tarih Kayması Çözüldü | TEFAS Excel Doğrudan Okuma | V17.7")

FUND_KINDS = ("YAT", "EMK", "BYF", "KAT", "")
DEFAULT_FUND_KIND = "YAT"
LOOKBACK_CALENDAR_DAYS = 45
TARGET_TRADING_DAYS = 10
MIN_ROLLING_DAYS = 5
MAX_WORKERS = 3
GITHUB_FALLBACK_URL = "https://github.com/tlgssk/kgdm3-fon-analiz/raw/refs/heads/main/Menkul_Kiymet_Yatirim_Fonlari_EXCEL_Tum_Veri.xlsx"

MOMENTUM_WEIGHTS = {"return": 0.30, "sharpe": 0.25, "cumulative": 0.25, "drawdown": 0.20}
SECURITY_SCALE = {"aum": 20.0, "investor": 20.0, "aum_flow": 8.0, "investor_change": 6.0}
Z_LIMIT = 2.5
BIST30_BONUS = 5.0
EMA_DECAY = 0.65

COLOR_NAVY, COLOR_GREEN, COLOR_RED, COLOR_YELLOW, COLOR_WHITE = "1F4E79", "008000", "FF0000", "B8860B", "FFFFFF"
COLOR_LIGHT_GREEN, COLOR_LIGHT_YELLOW, COLOR_LIGHT_RED = "E2F0D9", "FFF2CC", "FCE4D6"

# ============================================================
# YARDIMCI FONKSİYONLAR
# ============================================================
def clamp(value, low, high): return max(low, min(high, value))
def safe_float(value, default=0.0):
    try:
        if value is None: return default
        n = float(value)
        return default if pd.isna(n) else n
    except: return default
def optional_float(value):
    try:
        if value is None or (isinstance(value, str) and not value.strip()): return None
        n = float(value)
        return None if pd.isna(n) else n
    except: return None
def normalize_date_key(value):
    try:
        ts = pd.to_datetime(value, errors="coerce")
        return ts.strftime("%Y-%m-%d") if not pd.isna(ts) else None
    except: return None
def parse_number(value):
    if value is None or isinstance(value, bool): return None
    if isinstance(value, (int, float)): return None if pd.isna(value) else float(value)
    text = str(value).replace("₺", "").replace("TL", "").replace("%", "").replace(" ", "").strip()
    if not text: return None
    if "," in text and "." in text: text = text.replace(".", "").replace(",", ".") if text.rfind(",") > text.rfind(".") else text.replace(",", "")
    elif "," in text: text = text.replace(",", ".")
    elif "." in text and re.match(r"^-?\d{1,3}(\.\d{3})+$", text): text = text.replace(".", "")
    try: return float(text)
    except: return None
def normalize_fund_code(value):
    code = str(value).strip().upper()
    return code[:-2] if code.endswith(".0") else code

def calculate_compounded_return(returns):
    clean = [parse_number(v) for v in returns if v is not None]
    if not clean: return 0.0
    growth = 1.0
    for r in clean: growth *= 1.0 + r / 100.0
    return (growth - 1.0) * 100.0

def calculate_max_drawdown(prices):
    if not prices or len(prices) < 2: return 0.0
    peak, max_dd = safe_float(prices[0]), 0.0
    for price in prices:
        p = safe_float(price)
        if p <= 0: continue
        if p > peak: peak = p
        if peak > 0:
            dd = (p / peak - 1.0) * 100.0
            if dd < max_dd: max_dd = dd
    return max_dd

def zscore(values):
    clean = [optional_float(v) for v in values]
    valid = [v for v in clean if v is not None]
    if len(valid) < 2: return [0.0 if v is not None else None for v in clean]
    mean_v = sum(valid) / len(valid)
    std = (sum((x - mean_v) ** 2 for x in valid) / len(valid)) ** 0.5
    if std <= 1e-12: return [0.0 if v is not None else None for v in clean]
    out = []
    for v in clean:
        if v is not None: out.append(clamp((v - mean_v) / std, -Z_LIMIT, Z_LIMIT))
        else: out.append(None)
    return out

def detect_data_anomalies(prices: List[float], returns: List[float]) -> tuple:
    if not prices or len(prices) < 2: return False, ""
    if any(p <= 0 for p in prices if p is not None): return True, "Negatif/sıfır fiyat"
    if any(r > 20.0 or r < -20.0 for r in returns if r is not None): return True, "Günlük %20+ fiyat değişimi"
    return False, ""

# ============================================================
# ÇOK KATMANLI VERİ ÇEKME (API'LER VE EXCEL)
# ============================================================
@dataclass
class SourceStatus:
    source: str; attempted: bool = False; ok: bool = False; status_code: Optional[int] = None
    error_type: str = ""; message: str = ""; elapsed_ms: Optional[int] = None

def fetch_tier0_uploaded_excel(code: str, global_df: pd.DataFrame):
    status = SourceStatus("0. Hat: Yüklenen TEFAS Exceli", attempted=True)
    if global_df is None or global_df.empty: return None, status
    try:
        df = global_df[global_df["Fon Kodu"] == code].copy()
        if not df.empty:
            df["date"] = pd.to_datetime(df["Tarih"], errors="coerce", dayfirst=True)
            df["price"] = df["Fiyat"].apply(parse_number)
            df["aum"] = df["Portföy Büyüklüğü"].apply(parse_number) if "Portföy Büyüklüğü" in df.columns else None
            df["investors"] = df["Kişi Sayısı"].apply(parse_number) if "Kişi Sayısı" in df.columns else None
            df = df.dropna(subset=["date", "price"])[df["price"] > 0].sort_values("date").drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)
            if len(df) >= 2: status.ok = True; return df, status
    except Exception as e: status.message = str(e)[:100]
    return None, status

def fetch_tier1_tefascrawler(code: str, start, end):
    status = SourceStatus("1. Hat: tefas-crawler", attempted=True)
    if not HAS_TEFAS_CRAWLER:
        status.message = "tefas paketi kurulu değil"
        return None, status
    try:
        crawler = TefasCrawler()
        df = crawler.fetch(start=start.strftime("%Y-%m-%d"), end=end.strftime("%Y-%m-%d"), name=code)
        if df is not None and not df.empty:
            date_col = next((c for c in ["date", "TARIH", "Tarih"] if c in df.columns), None)
            price_col = next((c for c in ["price", "FIYAT", "Fiyat"] if c in df.columns), None)
            if date_col and price_col:
                df["date"] = pd.to_datetime(df[date_col], errors="coerce", dayfirst=True)
                df["price"] = df[price_col].apply(parse_number)
                df["aum"] = df["market_cap"].apply(parse_number) if "market_cap" in df.columns else None
                df["investors"] = df["number_of_investors"].apply(parse_number) if "number_of_investors" in df.columns else None
                df = df.dropna(subset=["date", "price"])[df["price"] > 0].sort_values("date").drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)
                if len(df) >= 2: status.ok = True; return df, status
    except Exception as e: status.message = str(e)[:100]
    return None, status

def fetch_tier2_tefas_direct_api(code: str, start, end, fund_kind: Optional[str] = None):
    status = SourceStatus("2. Hat: TEFAS Direct API", attempted=True)
    url = "https://www.tefas.gov.tr/api/DB/BindHistoryInfo"
    headers = {"X-Requested-With": "XMLHttpRequest", "Origin": "https://www.tefas.gov.tr"}
    for kind in ([fund_kind] if fund_kind in FUND_KINDS and fund_kind else list(FUND_KINDS)):
        payload = {"fontip": kind or "YAT", "fonkod": code, "bastarih": start.strftime("%d.%m.%Y"), "bittarih": end.strftime("%d.%m.%Y")}
        try:
            res = HTTP.post(url, data=payload, headers=headers, timeout=HTTP_TIMEOUT)
            status.status_code = res.status_code
            if res.status_code != 200: continue
            data = res.json().get("data", [])
            if not data: continue
            df = pd.DataFrame(data)
            df["date"] = pd.to_datetime(df["TARIH"], unit="ms", errors="coerce")
            df["price"] = df["FIYAT"].apply(parse_number)
            df["aum"] = df["PORTFOYBUYUKLUK"].apply(parse_number) if "PORTFOYBUYUKLUK" in df.columns else None
            df["investors"] = df["KISISAYISI"].apply(parse_number) if "KISISAYISI" in df.columns else None
            df = df.dropna(subset=["date", "price"])[df["price"] > 0].sort_values("date").drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)
            if len(df) >= 2: status.ok = True; return df, status
        except Exception as e: status.message = str(e)[:100]
    return None, status

def fetch_tier3_isyatirim(code: str, start, end):
    status = SourceStatus("3. Hat: İş Yatırım", attempted=True)
    url = "https://www.isyatirim.com.tr/_layouts/15/IsYatirim.Website/Common/Data.aspx/YatirimFonGecmisGetiri"
    params = {"fonKod": code, "baslangic": start.strftime("%d-%m-%Y"), "bitis": end.strftime("%d-%m-%Y")}
    try:
        res = HTTP.get(url, params=params, headers={"Accept": "application/json"}, timeout=HTTP_TIMEOUT)
        status.status_code = res.status_code
        if res.status_code != 200: return None, status
        df = pd.DataFrame(res.json().get("value", []))
        df["date"] = pd.to_datetime(df["Tarih"], dayfirst=True, errors="coerce")
        df["price"] = df["Fiyat"].apply(parse_number)
        df["aum"], df["investors"] = None, None
        df = df.dropna(subset=["date", "price"])[df["price"] > 0].sort_values("date").drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)
        if len(df) >= 2: status.ok = True; return df, status
    except Exception as e: status.message = str(e)[:100]
    return None, status

def get_fund_series(fund_code: str, fund_kind: Optional[str] = None, global_tefas_df=None):
    code = normalize_fund_code(fund_code)
    end = dt.datetime.now()
    start = end - dt.timedelta(days=LOOKBACK_CALENDAR_DAYS)
    statuses = []

    df0, s0 = fetch_tier0_uploaded_excel(code, global_tefas_df)
    if df0 is not None:
        statuses.append(s0)
        return df0, s0.source, statuses

    df1, s1 = fetch_tier1_tefascrawler(code, start, end)
    statuses.append(s1)
    if df1 is not None: return df1, s1.source, statuses

    df2, s2 = fetch_tier2_tefas_direct_api(code, start, end, fund_kind)
    statuses.append(s2)
    if df2 is not None: return df2, s2.source, statuses

    df3, s3 = fetch_tier3_isyatirim(code, start, end)
    statuses.append(s3)
    if df3 is not None: return df3, s3.source, statuses

    return None, "YOK", statuses

def fetch_fund_structural_data(fund_code: str, fund_title: Optional[str] = None) -> dict:
    structural = {"investment_area": "-"}
    t_upper = (fund_title or "").upper()
    if "PARA PİYASASI" in t_upper or "PPF" in t_upper: structural["investment_area"] = "Para Piyasası"
    elif "ALTIN" in t_upper or "GÜMÜŞ" in t_upper or "KIYMETLİ" in t_upper: structural["investment_area"] = "Kıymetli Maden"
    elif "YABANCI TEKNOLOJİ" in t_upper: structural["investment_area"] = "Hisse Senedi (Yabancı Teknoloji)"
    elif "YABANCI" in t_upper: structural["investment_area"] = "Yabancı Hisse/Fon"
    elif "HİSSE" in t_upper: structural["investment_area"] = "Hisse Senedi"
    elif "DEĞİŞKEN" in t_upper or "KARMA" in t_upper: structural["investment_area"] = "Karma / Değişken"
    return structural

def compute_fund_metrics(series: pd.DataFrame, fund_code: str, fund_title: Optional[str] = None):
    if series is None or len(series) < 2: return None
    df = series.copy().sort_values("date").drop_duplicates(subset=["date"], keep="last").reset_index(drop=True)
    df["date_key"] = df["date"].apply(normalize_date_key)
    df = df.dropna(subset=["date_key", "price"]).copy()
    df["price"] = df["price"].apply(optional_float)
    df = df[df["price"].notna() & (df["price"] > 0)].reset_index(drop=True)
    if len(df) < 2: return None

    prices = df["price"].tolist()
    date_keys_all = df["date_key"].tolist()
    aums = [optional_float(v) for v in df["aum"].tolist()] if "aum" in df.columns else [None] * len(df)
    invs = [optional_float(v) for v in df["investors"].tolist()] if "investors" in df.columns else [None] * len(df)

    aum_pct_changes = []
    for i in range(1, len(aums)):
        if aums[i] is not None and aums[i - 1] is not None and aums[i - 1] > 0:
            aum_pct_changes.append(((aums[i] / aums[i - 1]) - 1) * 100)
        else:
            aum_pct_changes.append(None)

    rets = []
    return_dates = []
    for i in range(1, len(prices)):
        if prices[i - 1] and prices[i - 1] > 0 and prices[i] and prices[i] > 0:
            rets.append((prices[i] / prices[i - 1] - 1.0) * 100.0)
            return_dates.append(date_keys_all[i])

    if not rets: return None

    has_anomaly, anomaly_reason = detect_data_anomalies(prices, rets)
    struct = fetch_fund_structural_data(fund_code, fund_title)
    
    aum_last = next((v for v in reversed(aums) if v is not None and v > 0), None)
    inv_last = next((v for v in reversed(invs) if v is not None and v >= 0), None)
    aum_first = next((v for v in aums if v is not None and v > 0), None)
    inv_first = next((v for v in invs if v is not None and v > 0), None)

    aum_change = ((aum_last / aum_first) - 1.0) * 100.0 if aum_last and aum_first else None
    inv_change = ((inv_last / inv_first) - 1.0) * 100.0 if inv_last is not None and inv_first else None

    mean_ret = sum(rets) / len(rets)
    vol = (sum((r - mean_ret) ** 2 for r in rets) / len(rets)) ** 0.5 if rets else 0.0

    return {
        "code": fund_code, "n_days": len(rets),
        "dates": return_dates, "daily_returns": rets, "prices": prices, "price_map": dict(zip(date_keys_all, prices)),
        "aum": aum_last, "investors": int(inv_last) if inv_last is not None else None,
        "aum_change": aum_change, "inv_change": inv_change,
        "aum_history_pct": aum_pct_changes[-30:],
        "volatility": vol,
        "fund_title": fund_title or "-", "has_anomaly": has_anomaly, "anomaly_reason": anomaly_reason,
        **struct,
    }

def fetch_and_compute_one_fund(code: str, meta_map: dict, global_tefas_df=None):
    meta = meta_map.get(code, {})
    series, source, statuses = get_fund_series(code, meta.get("kind", DEFAULT_FUND_KIND), global_tefas_df)
    metrics = compute_fund_metrics(series, code, meta.get("title"))
    if metrics is None: return code, None, source, statuses
    
    if meta.get("area"): metrics["investment_area"] = meta.get("area")
        
    metrics["source"] = source
    metrics["source_statuses"] = [asdict(s) for s in statuses]
    return code, metrics, source, statuses

# ============================================================
# GERÇEK MOMENTUM VE KARAR HESAPLAMALARI
# ============================================================
def calculate_real_trend_scores(funds: List[dict]):
    if not funds: return
    window = TARGET_TRADING_DAYS
    for f in funds:
        rets = f["daily_returns"][-window:] if f["daily_returns"] else []
        prices = f["prices"][-(window + 1):] if f["prices"] else []
        if rets:
            mean_r = sum(rets) / len(rets)
            vol = (sum((x - mean_r) ** 2 for x in rets) / len(rets)) ** 0.5
        else:
            mean_r, vol = 0.0, 0.0
        cum = (prices[-1] / prices[0] - 1.0) * 100.0 if len(prices) > 1 and prices[0] > 0 else 0.0
        dd = calculate_max_drawdown(prices) if prices else 0.0
        f["_mean_return"] = mean_r
        f["_sharpe"] = mean_r / vol if vol > 1e-12 else 0.0
        f["_cumulative"] = cum
        f["_maxdd"] = dd

    z_mean = zscore([f["_mean_return"] for f in funds])
    z_sharpe = zscore([f["_sharpe"] for f in funds])
    z_cum = zscore([f["_cumulative"] for f in funds])
    z_dd = zscore([-f["_maxdd"] for f in funds])
    z_aum = zscore([f.get("aum") for f in funds])
    z_inv = zscore([f.get("investors") for f in funds])
    z_aum_flow = zscore([f.get("aum_change") for f in funds])
    z_inv_flow = zscore([f.get("inv_change") for f in funds])

    for i, f in enumerate(funds):
        wz = (MOMENTUM_WEIGHTS["return"] * (z_mean[i] or 0.0) + MOMENTUM_WEIGHTS["sharpe"] * (z_sharpe[i] or 0.0) + MOMENTUM_WEIGHTS["cumulative"] * (z_cum[i] or 0.0) + MOMENTUM_WEIGHTS["drawdown"] * (z_dd[i] or 0.0))
        momentum = clamp(50.0 + 20.0 * wz, 0.0, 100.0)

        sec = 50.0
        sec += SECURITY_SCALE["aum"] * 0.30 * (z_aum[i] or 0.0)
        sec += SECURITY_SCALE["investor"] * 0.25 * (z_inv[i] or 0.0)
        if f.get("aum_change") is not None: sec += SECURITY_SCALE["aum_flow"] * (z_aum_flow[i] or 0.0)
        if f.get("inv_change") is not None: sec += SECURITY_SCALE["investor_change"] * (z_inv_flow[i] or 0.0)
        security = clamp(sec, 0.0, 100.0)

        f["market_momentum"] = round(momentum, 1)
        f["security_score"] = round(security, 1)

    for f in funds:
        rets_all = f["daily_returns"]
        prices_all = f["prices"]
        
        running = [None] * min(len(rets_all), MIN_ROLLING_DAYS - 1)
        
        for d in range(MIN_ROLLING_DAYS, len(rets_all) + 1):
            r_slice = rets_all[d - MIN_ROLLING_DAYS:d]
            p_slice = prices_all[d - MIN_ROLLING_DAYS:d + 1]
            mr = sum(r_slice) / len(r_slice)
            vol = (sum((x - mr) ** 2 for x in r_slice) / len(r_slice)) ** 0.5
            cum = calculate_compounded_return(r_slice)
            local_z = clamp(mr / vol, -Z_LIMIT, Z_LIMIT) if vol > 1e-12 else 0.0
            score = clamp(50.0 + 15.0 * local_z + clamp(cum, -20, 20) * 0.5, 0.0, 100.0)
            running.append(round(score, 1))
        f["running_trend_hybrid"] = running

def finalize_decisions(funds: List[dict], yatirim_vadesi: str, macro_diagnostics: dict):
    hz = FonalistEngine.weights_for_horizon(yatirim_vadesi)
    engine = FonalistEngine(base_threshold=hz["base_threshold"], z_score_threshold=-2.5)

    bist_proxy = macro_diagnostics.get("bist_proxy", 0.0)
    us10y_proxy = macro_diagnostics.get("us10y_proxy", 0.0)
    brent_proxy = macro_diagnostics.get("brent_proxy", 0.0)
    current_regime = engine.detect_macro_regime(bist_proxy, us10y_proxy, brent_proxy)

    for f in funds:
        f["decision_history"] = []
        mom = f.get("market_momentum", 50.0)
        sec = f.get("security_score", 50.0)
        sent = 50.0 

        live_chg = get_live_proxy_change(f.get("code"), f.get("investment_area", "")) * hz["proxy_scale"]
        f["live_proxy_chg"] = round(live_chg, 2)
        f["macro_regime"] = current_regime.value
        f["yatirim_vadesi"] = yatirim_vadesi

        trend_scores = f.get("running_trend_hybrid", [])
        dates = f.get("dates", [])
        
        temp_history = []
        for i in range(max(0, len(trend_scores) - 5), len(trend_scores)):
            current_date = dates[i] if i < len(dates) else f"T-{len(trend_scores)-i}"
            current_score = trend_scores[i]
            
            if current_score is None: continue
            
            hist_scores_up_to_now = temp_history[::-1] + [current_score]

            fund_metrics = FundMetrics(
                code=f.get("code"), name=f.get("fund_title", ""), category=f.get("investment_area", ""),
                current_score=current_score, history_scores=hist_scores_up_to_now,
                volatility=safe_float(f.get("volatility", 1.0)), aum_change_pct=safe_float(f.get("aum_change")),
                investor_change_pct=safe_float(f.get("inv_change")), aum_history_pct=f.get("aum_history_pct", []),
                underlying_proxy_chg=live_chg if i == len(trend_scores) - 1 else 0.0
            )
            decision = engine.evaluate_fund(fund_metrics, current_regime)
            temp_history.append(current_score)
            
            f["decision_history"].append({
                "date": current_date, "score": decision["skor"], "decision": decision["sinyal"].value,
                "dynamic_threshold": decision["esik"], "action_reason": decision["gerekce"]
            })
            
        f["decision_history"].reverse() 
        if f["decision_history"]:
            latest = f["decision_history"][0]
            f["decision_score"] = round(latest["score"], 1)
            f["karar"] = latest["decision"]
            f["dynamic_threshold"] = round(latest["dynamic_threshold"], 1)
            f["action_reason"] = latest["action_reason"]
            
            if len(f["decision_history"]) >= 2:
                d0, d1 = f["decision_history"][0]["decision"], f["decision_history"][1]["decision"]
                f["urgent_sell_2day"] = (d0 in [ActionSignal.STOP_LOSS_30.value, ActionSignal.EMERGENCY_EXIT.value] and d1 in [ActionSignal.STOP_LOSS_30.value, ActionSignal.EMERGENCY_EXIT.value])
            else:
                f["urgent_sell_2day"] = False

# ============================================================
# EXCEL İHRACAT FONKSİYONU
# ============================================================
def create_excel_output(wb, ws_list, all_funds, failed_codes, recent_5_formatted):
    if "FONALIST_Analiz" in wb.sheetnames: del wb["FONALIST_Analiz"]
    ws_scores = wb.create_sheet(title="FONALIST_Analiz")

    headers = [
        "Fon Kodu", "Fon Adı", "Yatırım Alanı", "Veri Kaynağı", "Makro Rejim", "Yatırım Vadesi", "Anomali Uyarısı",
        "Volatilite (%)", "AUM Değişim (%)", "AUM (₺)", "Yatırımcı"
    ]
    for _, fmt_date in recent_5_formatted:
        headers.append(f"{fmt_date} Skor")
        headers.append(f"{fmt_date} Karar")

    ws_scores.append(headers)
    
    header_fill = PatternFill(start_color=COLOR_NAVY, end_color=COLOR_NAVY, fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color=COLOR_WHITE)
    for cell in ws_scores[1]:
        cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for item in all_funds:
        anomali_msg = item.get("anomaly_reason") if item.get("has_anomaly") else "Yok"
        row_data = [
            item["code"], item.get("fund_title") or "-", item.get("investment_area") or "-", item.get("source", "-"),
            item.get("macro_regime", "-"), item.get("yatirim_vadesi", "-"), anomali_msg, 
            item.get("volatility"), round(safe_float(item.get("aum_change")), 2) if item.get("aum_change") is not None else None, 
            round(safe_float(item.get("aum")), 0) if item.get("aum") is not None else None, item.get("investors")
        ]
        
        hist_map = {h.get("date"): h for h in item.get("decision_history", [])}
        for orig_date, _ in recent_5_formatted:
            h = hist_map.get(orig_date)
            if h:
                row_data.append(h.get("score"))
                row_data.append(h.get("decision"))
            else:
                row_data.extend(["-", "-"])
                
        ws_scores.append(row_data)

    if failed_codes:
        ws_scores.append([])
        ws_scores.append(["VERİ ALINAMAYAN FONLAR (Gerçek veri bulunamadı)"])
        for code in failed_codes: ws_scores.append([code])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ============================================================
# UI (SIDEBAR & MAIN)
# ============================================================
st.sidebar.header("⚙️ Analiz & Filtre Kriterleri")
yatirim_vadesi = st.sidebar.radio("Vade Seçin", ["Kısa Vade", "Orta Vade", "Uzun Vade"], index=1)
st.sidebar.markdown("---")
st.sidebar.info("💡 **Z-Score Anomali Motoru Aktif:** Fon likidite krizlerini, gün içi varlık şoklarını hesaplar ve 5 günlük geçmişi haritalandırır.")
SHOW_DIAGNOSTICS = st.sidebar.checkbox("Kaynak tanılama bilgisini göster", value=True)

if not HAS_TEFAS_CRAWLER:
    st.sidebar.warning("⚠️ `tefas` paketi kurulu değil — Doğrudan TEFAS Excel Yüklemeye veya API Hatlarına Düşülecek.")

st.markdown("### 📥 Veri Kaynağı Seçimi")
col_upload, col_github, col_manual = st.columns(3)
wb = None
ws_list = None
manuel_req_codes = []
global_tefas_df = None

with col_upload:
    uploaded_file = st.file_uploader("Bilgisayardan Excel Yükle (Fon Listesi veya TEFAS Dosyası)", type=["xlsx"])
    if uploaded_file is not None:
        file_bytes = uploaded_file.getvalue()
        try:
            temp_df = pd.read_excel(io.BytesIO(file_bytes))
            temp_df.columns = temp_df.columns.str.strip()
            if "Tarih" in temp_df.columns and "Fon Kodu" in temp_df.columns and "Fiyat" in temp_df.columns:
                global_tefas_df = temp_df
                st.success("✅ TEFAS Tarihsel Veri Excel'i başarıyla algılandı. API ve Crawler atlanıp fiyatlar bu dosyadan okunacak!")
        except:
            pass

        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
            ws_list = wb["Fon_Listesi"] if "Fon_Listesi" in wb.sheetnames else wb.active
        except Exception as exc: 
            st.error(f"Excel Listesi okuma hatası: {exc}")

with col_github:
    st.markdown("<br><br>", unsafe_allow_html=True)
    if st.button("🚀 GitHub'dan Çek ve Analiz Et", use_container_width=True):
        url = GITHUB_FALLBACK_URL
        try:
            res = requests.get(url)
            if res.status_code == 200:
                wb = openpyxl.load_workbook(io.BytesIO(res.content))
                ws_list = wb["Fon_Listesi"] if "Fon_Listesi" in wb.sheetnames else wb.active
                st.success("✅ Liste GitHub'dan çekildi.")
            else:
                st.error(f"GitHub'dan çekilemedi. Hata kodu: {res.status_code}")
        except Exception as e:
            st.error(f"Bağlantı hatası: {e}")

with col_manual:
    manuel_fonlar = st.text_area("✍️ Manuel Fon Kodları", placeholder="Örn: AFS, THF, KZL, BCO, IDI, KHA\n(Virgülle ayırın)")
    if manuel_fonlar:
        manuel_req_codes = [normalize_fund_code(f.strip()) for f in manuel_fonlar.split(",") if f.strip()]

req_codes = []
excel_meta_override = {} 

if manuel_req_codes:
    req_codes = manuel_req_codes
    if not wb:
        wb = openpyxl.Workbook()
        ws_list = wb.active
        ws_list.title = "Fon_Listesi"
        ws_list.append(["Fon Kodu"])
    for code in req_codes: ws_list.append([code])
elif wb is not None and ws_list is not None:
    is_tefas_dump = "Tarih" in [str(c.value).strip() for c in ws_list[1] if c.value]
    
    if is_tefas_dump and global_tefas_df is not None:
        req_codes = list(global_tefas_df["Fon Kodu"].dropna().unique())
        for code in req_codes:
            f_title = global_tefas_df[global_tefas_df["Fon Kodu"] == code]["Fon Adı"].iloc[0] if "Fon Adı" in global_tefas_df.columns else ""
            excel_meta_override[code] = {"title": str(f_title).strip(), "area": ""}
    else:
        for r in ws_list.iter_rows(min_row=2):
            if r and r[0].value:
                code = normalize_fund_code(r[0].value)
                if code:
                    req_codes.append(code)
                    title_val = str(r[1].value).strip() if len(r) > 1 and r[1].value and str(r[1].value).strip() != "None" else ""
                    area_val = str(r[2].value).strip() if len(r) > 2 and r[2].value and str(r[2].value).strip() != "None" else ""
                    if title_val or area_val:
                        excel_meta_override[code] = {"title": title_val, "area": area_val}

req_codes = list(dict.fromkeys(filter(None, req_codes)))

if st.button("🚀 FONALİST Motorunu Çalıştır", use_container_width=True, type="primary"):
    if not req_codes:
        st.error("Lütfen bir Excel yükleyin, GitHub'dan çekin veya manuel fon kodu girin.")
        st.stop()

    with st.spinner("🔄 Gerçek veriler çok katmanlı yapıyla çekiliyor..."):
        meta_map = {}
        for c in req_codes:
            meta_map[c] = {
                "title": excel_meta_override.get(c, {}).get("title", c),
                "area": excel_meta_override.get(c, {}).get("area", "")
            }
            
        calc_funds, failed_codes, all_statuses = [], [], []

        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as exe:
            futs = {exe.submit(fetch_and_compute_one_fund, c, meta_map, global_tefas_df): c for c in req_codes}
            for fut in concurrent.futures.as_completed(futs):
                code = futs[fut]
                try: _, met, src, statuses = fut.result()
                except Exception: met, statuses = None, []
                
                if met: calc_funds.append(met)
                else: failed_codes.append(code)
                for s in statuses: all_statuses.append({"Fon": code, **asdict(s)})
        
        if failed_codes:
            st.warning(f"⚠️ {len(failed_codes)} fon için GERÇEK veri alınamadı ve hesaplama dışı bırakıldı: {', '.join(failed_codes)}")

        if not calc_funds:
            st.error("Veri alınamadı. İşlem durduruldu.")
            st.stop()

        eligible = [f for f in calc_funds if f.get("n_days", 0) >= MIN_ROLLING_DAYS]
        thin_data = [f for f in calc_funds if f.get("n_days", 0) < MIN_ROLLING_DAYS]
        all_calc = eligible + thin_data

        calculate_real_trend_scores(all_calc)

        macro_diagnostics = {
            "bist_proxy": get_live_proxy_change("THF", "Hisse Senedi"),
            "us10y_proxy": get_live_proxy_change("US10Y", "Tahvil"),
            "brent_proxy": get_live_proxy_change("BRENT", "Emtia"),
        }

        finalize_decisions(all_calc, yatirim_vadesi, macro_diagnostics)
        
        all_dates = set()
        for item in all_calc:
            for h in item.get("decision_history", []):
                all_dates.add(h.get("date"))
        
        recent_5_dates = sorted(list(all_dates), reverse=True)[:5]
        recent_5_formatted = []
        for d_str in recent_5_dates:
            try: fmt = dt.datetime.strptime(d_str, "%Y-%m-%d").strftime("%d.%m.%Y")
            except: fmt = d_str
            recent_5_formatted.append((d_str, fmt))

        if not wb:
            wb = openpyxl.Workbook()
            ws_list = wb.active
            
        output = create_excel_output(wb, ws_list, all_calc, failed_codes, recent_5_formatted)

    g_al_fonlar = []
    a_sat_fonlar = []
    
    for item in all_calc:
        if item.get("urgent_sell_2day"):
            a_sat_fonlar.append(item["code"])
        else:
            hist = item.get("decision_history", [])
            if len(hist) >= 2 and hist[0].get("decision") == ActionSignal.STRONG_BUY.value and hist[1].get("decision") == ActionSignal.STRONG_BUY.value:
                g_al_fonlar.append(item["code"])

    if g_al_fonlar: st.success(f"🚀 **2 GÜN ÜST ÜSTE GÜÇLÜ AL SİNYALİ ÜRETENLER:** {', '.join(g_al_fonlar)}")
    if a_sat_fonlar: st.error(f"🚨 **2 GÜN ÜST Üste ACİL SAT / TASFİYE SİNYALİ ÜRETENLER:** {', '.join(a_sat_fonlar)}")

    st.subheader(f"📈 FONALİST 5 Günlük Sinyal Matrisi")
    st.caption(f"Makro Rejim: **{macro_diagnostics and all_calc[0].get('macro_regime', '-') if all_calc else '-'}** | Vade: **{yatirim_vadesi}**")
    
    display_rows = []
    for item in all_calc:
        row_dict = {
            "Fon Kodu": item["code"],
            "Fon Adı": item.get("fund_title", "-"),
            "Yatırım Alanı": item.get("investment_area", "-")
        }
        hist_map = {h.get("date"): h for h in item.get("decision_history", [])}
        for orig_date, fmt_date in recent_5_formatted:
            h = hist_map.get(orig_date)
            row_dict[f"{fmt_date} Skor"] = h.get("score") if h else "-"
            row_dict[f"{fmt_date} Karar"] = h.get("decision") if h else "-"
            
        row_dict["Anomali/Gerekçe"] = item.get("action_reason", "-")
        display_rows.append(row_dict)

    df_display = pd.DataFrame(display_rows)

    def color_cells(value):
        text = str(value).upper()
        if "GÜÇLÜ AL" in text or "TUT" in text: return "color: #008000; font-weight: bold;"
        if "İZLE" in text: return "color: #B8860B; font-weight: bold;"
        if "TASFİYE" in text or "SAT" in text: return "color: #FF0000; font-weight: bold;"
        return ""

    styled_df = df_display.style.map(color_cells) if hasattr(df_display.style, "map") else df_display.style.applymap(color_cells)
    st.dataframe(styled_df, use_container_width=True, hide_index=True)
    
    st.download_button(label="📥 FONALİST Excel İndir", data=output, file_name="fonalist_analiz_v17_7.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    if SHOW_DIAGNOSTICS and all_statuses:
        st.subheader("🔎 Veri Kaynağı Tanılaması")
        df_diag = pd.DataFrame(all_statuses)
        st.dataframe(df_diag, use_container_width=True, hide_index=True)
